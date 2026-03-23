"""수집 결과 엑셀 보고서 + 미리보기 HTML 생성"""
import json, shutil, random
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent.parent
CLASSIFIED_DIR = BASE_DIR / "classified"
METADATA_DIR = BASE_DIR / "metadata"
RESULT_FILE = METADATA_DIR / "classification_result.json"

STYLE_NAMES_KR = {
    "modern": "모던", "nordic": "북유럽", "classic": "클래식",
    "natural": "내추럴", "industrial": "인더스트리얼", "luxury": "럭셔리"
}
STYLES = ["modern", "nordic", "classic", "natural", "industrial", "luxury"]

# ── Load data ──
data = json.loads(RESULT_FILE.read_text(encoding="utf-8"))
images = data.get("images", {})

# ── Aggregate stats ──
style_data = {}
low_conf_list = []
for name, info in images.items():
    s = info.get("style", "unknown")
    c = info.get("confidence", 0)
    r = info.get("reason", "")
    style_data.setdefault(s, []).append({"name": name, "confidence": c, "reason": r})
    if c < 0.6:
        low_conf_list.append({"name": name, "style": s, "confidence": c, "reason": r})

# ═══════════════════════════════════════
# 1. Excel Report
# ═══════════════════════════════════════
wb = Workbook()

# --- Sheet 1: Summary ---
ws = wb.active
ws.title = "스타일별 요약"

header_font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="2E75B6")
sub_font = Font(name="Arial", bold=True, size=10)
sub_fill = PatternFill("solid", fgColor="D6E4F0")
data_font = Font(name="Arial", size=10)
border = Border(
    left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"), bottom=Side(style="thin", color="CCCCCC")
)
center = Alignment(horizontal="center", vertical="center")

ws.merge_cells("A1:E1")
ws["A1"] = "한국형 주방 이미지 수집 & 분류 결과 보고서"
ws["A1"].font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
ws["A1"].fill = PatternFill("solid", fgColor="1B4F72")
ws["A1"].alignment = center

ws["A2"] = f"생성일: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
ws["A2"].font = Font(name="Arial", size=9, italic=True, color="666666")

headers = ["스타일", "스타일(영문)", "이미지 수", "평균 신뢰도", "비율"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border

total_count = len(images)
for row_idx, s in enumerate(STYLES, 5):
    items = style_data.get(s, [])
    count = len(items)
    avg_conf = sum(i["confidence"] for i in items) / max(len(items), 1)

    ws.cell(row=row_idx, column=1, value=STYLE_NAMES_KR.get(s, s)).font = data_font
    ws.cell(row=row_idx, column=2, value=s).font = data_font
    ws.cell(row=row_idx, column=3, value=count).font = data_font
    ws.cell(row=row_idx, column=3).alignment = center
    ws.cell(row=row_idx, column=4, value=round(avg_conf, 3)).font = data_font
    ws.cell(row=row_idx, column=4).number_format = "0.0%"
    ws.cell(row=row_idx, column=4).alignment = center
    if total_count > 0:
        ws.cell(row=row_idx, column=5).font = data_font
        ws.cell(row=row_idx, column=5).number_format = "0.0%"
        ws.cell(row=row_idx, column=5).alignment = center
        ws.cell(row=row_idx, column=5, value=f"=C{row_idx}/C11")

    for c in range(1, 6):
        ws.cell(row=row_idx, column=c).border = border

# Total row
total_row = 5 + len(STYLES)
ws.cell(row=total_row, column=1, value="합계").font = sub_font
ws.cell(row=total_row, column=1).fill = sub_fill
ws.cell(row=total_row, column=2, value="").fill = sub_fill
ws.cell(row=total_row, column=3, value=f"=SUM(C5:C{total_row-1})").font = sub_font
ws.cell(row=total_row, column=3).fill = sub_fill
ws.cell(row=total_row, column=3).alignment = center
ws.cell(row=total_row, column=4, value="").fill = sub_fill
ws.cell(row=total_row, column=5, value="").fill = sub_fill
for c in range(1, 6):
    ws.cell(row=total_row, column=c).border = border

ws.column_dimensions["A"].width = 16
ws.column_dimensions["B"].width = 16
ws.column_dimensions["C"].width = 14
ws.column_dimensions["D"].width = 14
ws.column_dimensions["E"].width = 12

# --- Sheet 2: Low confidence ---
ws2 = wb.create_sheet("낮은 신뢰도 (0.6 미만)")
headers2 = ["파일명", "분류 스타일", "신뢰도", "분류 근거"]
for col, h in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border

for row_idx, item in enumerate(sorted(low_conf_list, key=lambda x: x["confidence"]), 2):
    ws2.cell(row=row_idx, column=1, value=item["name"]).font = data_font
    ws2.cell(row=row_idx, column=2, value=STYLE_NAMES_KR.get(item["style"], item["style"])).font = data_font
    ws2.cell(row=row_idx, column=3, value=item["confidence"]).font = data_font
    ws2.cell(row=row_idx, column=3).number_format = "0.0%"
    ws2.cell(row=row_idx, column=4, value=item["reason"]).font = data_font
    for c in range(1, 5):
        ws2.cell(row=row_idx, column=c).border = border

ws2.column_dimensions["A"].width = 25
ws2.column_dimensions["B"].width = 16
ws2.column_dimensions["C"].width = 12
ws2.column_dimensions["D"].width = 50

# --- Sheet 3: Full list ---
ws3 = wb.create_sheet("전체 분류 목록")
headers3 = ["파일명", "스타일", "스타일(한글)", "신뢰도", "분류 근거", "분류 시각"]
for col, h in enumerate(headers3, 1):
    cell = ws3.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border

for row_idx, (name, info) in enumerate(sorted(images.items()), 2):
    ws3.cell(row=row_idx, column=1, value=name).font = data_font
    ws3.cell(row=row_idx, column=2, value=info.get("style", "")).font = data_font
    ws3.cell(row=row_idx, column=3, value=STYLE_NAMES_KR.get(info.get("style", ""), "")).font = data_font
    ws3.cell(row=row_idx, column=4, value=info.get("confidence", 0)).font = data_font
    ws3.cell(row=row_idx, column=4).number_format = "0.0%"
    ws3.cell(row=row_idx, column=5, value=info.get("reason", "")).font = data_font
    ws3.cell(row=row_idx, column=6, value=info.get("classified_at", "")).font = data_font
    for c in range(1, 7):
        ws3.cell(row=row_idx, column=c).border = border

ws3.column_dimensions["A"].width = 25
ws3.column_dimensions["B"].width = 14
ws3.column_dimensions["C"].width = 14
ws3.column_dimensions["D"].width = 12
ws3.column_dimensions["E"].width = 50
ws3.column_dimensions["F"].width = 22

excel_path = BASE_DIR / "summary.xlsx"
wb.save(str(excel_path))
print(f"✅ 엑셀 저장: {excel_path}")

# ═══════════════════════════════════════
# 2. Preview HTML
# ═══════════════════════════════════════
style_colors = {
    "modern": "#3498db", "nordic": "#2ecc71", "classic": "#9b59b6",
    "natural": "#e67e22", "industrial": "#7f8c8d", "luxury": "#f1c40f"
}

html_parts = ["""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<title>주방 이미지 스타일 분류 미리보기</title>
<style>
* { margin:0; padding:0; box-sizing:border-box; }
body { font-family:'Segoe UI',Arial,sans-serif; background:#f5f5f5; padding:30px; }
h1 { text-align:center; color:#1B4F72; margin-bottom:10px; font-size:24px; }
.subtitle { text-align:center; color:#666; margin-bottom:30px; font-size:14px; }
.stats { display:flex; justify-content:center; gap:15px; flex-wrap:wrap; margin-bottom:30px; }
.stat-card { background:white; border-radius:12px; padding:15px 25px; text-align:center; box-shadow:0 2px 8px rgba(0,0,0,0.1); min-width:120px; }
.stat-card .num { font-size:28px; font-weight:bold; }
.stat-card .label { font-size:12px; color:#888; margin-top:4px; }
.style-section { margin-bottom:40px; }
.style-header { display:flex; align-items:center; gap:12px; margin-bottom:15px; padding:10px 20px; border-radius:10px; color:white; }
.style-header h2 { font-size:18px; }
.style-header .count { margin-left:auto; font-size:14px; opacity:0.9; }
.grid { display:grid; grid-template-columns:repeat(auto-fill, minmax(200px, 1fr)); gap:12px; padding:0 10px; }
.card { background:white; border-radius:8px; overflow:hidden; box-shadow:0 2px 6px rgba(0,0,0,0.08); transition:transform 0.2s; }
.card:hover { transform:translateY(-3px); box-shadow:0 4px 12px rgba(0,0,0,0.15); }
.card img { width:100%; height:160px; object-fit:cover; }
.card .info { padding:8px 10px; }
.card .info .name { font-size:11px; color:#333; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }
.card .info .conf { font-size:11px; color:#888; }
.footer { text-align:center; color:#aaa; font-size:12px; margin-top:40px; padding-top:20px; border-top:1px solid #ddd; }
</style>
</head>
<body>
<h1>한국형 주방 이미지 스타일 분류 미리보기</h1>
"""]

html_parts.append(f'<p class="subtitle">생성일: {datetime.now().strftime("%Y-%m-%d %H:%M")} | 총 {total_count}장</p>')

# Stats cards
html_parts.append('<div class="stats">')
for s in STYLES:
    items = style_data.get(s, [])
    count = len(items)
    color = style_colors.get(s, "#333")
    html_parts.append(f'<div class="stat-card"><div class="num" style="color:{color}">{count}</div><div class="label">{STYLE_NAMES_KR[s]}</div></div>')
html_parts.append('</div>')

# Style sections with representative images
for s in STYLES:
    items = style_data.get(s, [])
    if not items:
        continue

    color = style_colors.get(s, "#333")
    count = len(items)
    avg_conf = sum(i["confidence"] for i in items) / max(len(items), 1)

    # Pick top 6 by confidence
    top_items = sorted(items, key=lambda x: x["confidence"], reverse=True)[:6]

    html_parts.append(f'<div class="style-section">')
    html_parts.append(f'<div class="style-header" style="background:{color}">')
    html_parts.append(f'<h2>{STYLE_NAMES_KR[s]} ({s})</h2>')
    html_parts.append(f'<span class="count">{count}장 | 평균 신뢰도 {avg_conf:.0%}</span>')
    html_parts.append(f'</div>')
    html_parts.append(f'<div class="grid">')

    for item in top_items:
        img_path = f"classified/{s}/{item['name']}"
        conf = item["confidence"]
        html_parts.append(f'''<div class="card">
<img src="{img_path}" alt="{item['name']}" loading="lazy" onerror="this.style.display='none'">
<div class="info"><div class="name">{item['name']}</div><div class="conf">신뢰도: {conf:.0%}</div></div>
</div>''')

    html_parts.append('</div></div>')

html_parts.append(f'<div class="footer">다담가구 AI 시스템팀 — {datetime.now().strftime("%Y.%m.%d")}</div>')
html_parts.append('</body></html>')

html_path = BASE_DIR / "preview.html"
html_path.write_text("\n".join(html_parts), encoding="utf-8")
print(f"✅ HTML 저장: {html_path}")
print("✅ 보고서 생성 완료!")
